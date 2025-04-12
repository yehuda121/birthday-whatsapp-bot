import datetime
import time
import re
import pywhatkit
from openpyxl import load_workbook
import os

def log(message):
    print(message)
    with open("C:\\Users\\Moked Kishla\\Desktop\\Birthdays\\log.txt", "a", encoding="utf-8") as log_file:
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_file.write(f"[{timestamp}] {message}\n")

def normalize_phone(raw_phone):
    if raw_phone is None:
        return None
    phone = str(raw_phone)
    phone = re.sub(r"[^\d]", "", phone)  
    if phone.startswith("972"):
        return "+" + phone
    elif phone.startswith("05"):
        return "+972" + phone[1:]
    elif phone.startswith("+972"):
        return phone
    return None

def extract_day_month(cell_value):
    if isinstance(cell_value, datetime.datetime):
        return cell_value.day, cell_value.month
    try:
        parsed = datetime.datetime.strptime(str(cell_value), "%d/%m")
        return parsed.day, parsed.month
    except:
        return None, None

wb = load_workbook("C:\\Users\\Moked Kishla\\Desktop\\Birthdays\\birthday.xlsx", data_only=True)
ws = wb.active

today = datetime.datetime.today()
today_day = today.day
today_month = today.month

sent_counter = 0
log(f"today: {today_day}/{today_month}")

for row in range(2, ws.max_row + 1):
    birth_cell = ws[f"C{row}"].value
    if not birth_cell:
        continue

    day, month = extract_day_month(birth_cell)
    # log(f"row {row} - cell date: {birth_cell} =>: {day}/{month}")
    name = ws[f"B{row}"].value or "no name"
    # log(name.strip())
    # print(name.strip())
    if day == today_day and month == today_month:
        # name = ws[f"B{row}"].value or "no name"
        raw_phone = ws[f"E{row}"].value
        phone_number = normalize_phone(raw_phone)

        if phone_number is None:
            log(f"phone number not valid {row}: {raw_phone}")
            continue

        if name.strip() == "שי":
            message = "יש לך יום הולדת אבל אל תעוף על עצמך ואל תיהיה כלכך זקן. תבוא לבקר את יהודה ותביא איתך בירה! 🍻"
        elif name.strip() == "אנוקה":
            message = "יש לך יום הולדת אבל אל תעוף על עצמך ואל תיהיה כלכך זקן. תבוא לבקר את יהודה ותביא איתך בירה! 🍻"
        else:
            message = f"שלום {name}! מזל טוב ליום הולדתך! 🎉 מיחידת העיר העתיקה."
        # print(message)
        

        try:
            pywhatkit.sendwhatmsg_instantly(phone_number, message, wait_time=10, tab_close=True)
            time.sleep(10)
            log(f" נשלחה הודעה ל: {name}")
            sent_counter += 1
        except Exception as e:
            log(f" שגיאה בשליחת הודעה ל-{name}: {e}")


if sent_counter == 0:
    log("no birthdays today!")
else:
    log(f"sended {sent_counter} messages.")